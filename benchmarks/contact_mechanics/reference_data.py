"""Loader for the paper's semi-analytical dataset (4TU repository).

Novikov et al. (2024), data at https://doi.org/10.4121/d77f1a2c-29ea-4572-ad72-e33ed8dc8d22
(``Data_GGGG_NovikovEtAl2024.xlsx``, md5 a0471b397331b0edcce4dadbd6272e7f).
Each row: figure label, variable name, units, then the values.

Known defects in the published file (verified element-wise):

* every ``delta`` row -- ``6 right``, ``9 & 10 right`` (both ``delta_25``
  and ``delta_27``) and ``14 right`` -- holds the *same* 2048-point curve,
  the Fig. 14 pre-nucleation profile (two patches, peaks 1.86 / 3.74 mm);
  the slip curves of Figs. 6, 9 and 10 are therefore not in the dataset;
* ``6 left`` ``Sigma_C_pre`` duplicates ``8 right`` (the inclined-fault
  curve: it satisfies ``Sigma_C_pre = Sigma_shear - Sigma_slip`` of
  ``8 left`` to 1e-13, and does not match Fig. 6's vertical-fault
  analytic), so Fig. 6's stress curve is missing too;
* the ``12`` rows are scrambled: ``-p_1`` holds y-coordinates while
  ``y_tilde_1``, ``-p_4`` and ``y_tilde_4`` are three copies of one
  pressure row, and pairs 2 and 3 duplicate each other -- only two of the
  four patch-boundary curves survive, under unreliable names.

The stress rows of Figs. 3, 4, 8, 9/10 (left), 14 (left) are genuine.
"""

from __future__ import annotations

from pathlib import Path

import numpy as np

DATA = Path(__file__).parent / "_data" / "Data_GGGG_NovikovEtAl2024.xlsx"
_cache: dict = {}


def load(figure: str, path=DATA) -> dict:
    """All variables of one (sub)figure: ``{name: np.ndarray}``."""
    key = (str(path), figure)
    if key not in _cache:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True)
        out = {}
        for row in wb[wb.sheetnames[0]].iter_rows(values_only=True):
            if str(row[0]) == figure and row[1] is not None:
                values = [v for v in row[3:] if v is not None]
                out[str(row[1])] = np.asarray(values, dtype=float)
        wb.close()
        if not out:
            raise KeyError(f"figure {figure!r} not in {path}")
        _cache[key] = out
    return _cache[key]
